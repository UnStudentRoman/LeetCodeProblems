import openpyxl
import pandas as pd
import re


if __name__ == '__main__':



    excel_template = input("Write excel full path:")
    sheet_value = input("Write sheet name:")


    # Load the Excel file
    workbook = openpyxl.load_workbook(rf"{excel_template}.xlsx")
    sheet = workbook[sheet_value]

    # Create an empty dictionary to store the data
    data_dict = {}
    last_header_value = None

    # Iterate through the rows in column B and extract data for each value
    for row in range(2, sheet.max_row + 1, 3):

        for header_col in range(3, len(sheet[row])+1):

            cell_value = sheet.cell(row=row, column=2).value
            if cell_value not in data_dict:
                data_dict[cell_value] = {}
            header_value = sheet.cell(row=row, column=header_col).value
            if header_value is None:
                header_value = last_header_value
            else:
                last_header_value = header_value
            if cell_value is not None:
                current_key = cell_value
                current_header = header_value
                current_header = current_header.replace(" ", "")
                if current_header not in data_dict[current_key]:
                    data_dict[current_key][current_header] = []
            # Gather data for the corresponding value in column B
            for row_offset in range(1, 3):
                data_cell = sheet.cell(row=row + row_offset, column=header_col).value
                if data_cell is not None:
                    data_cell = data_cell.rstrip(" ")
                    data_dict[current_key][current_header].append(data_cell)

    # Print the resulting dictionary
    # print(data_dict)

    for circuit in data_dict:
        for site in data_dict[circuit]:
            change_list = []
            pattern = r"^[A-Za-z]{4}_\d$"
            for slot in data_dict[circuit][site]:

                # Iterate through the input strings and check if they match the pattern
                if re.match(pattern, slot):
                    change_list.append(slot)
            for slot in data_dict[circuit][site]:
                if slot not in change_list:
                    change_list.append(slot)
            data_dict[circuit][site] = change_list

    # print(data_dict)

    df = pd.DataFrame(columns=["Site", "Circuit", "Slot", "BTB"])


    for circuit in data_dict:
        for site in data_dict[circuit]:
            # df['Site'] = site
            # df['Circuit'] = circuit
            for slot in data_dict[circuit][site]:
                if slot[0] == "B" or slot[0] == "C":
                    btb = slot
            slots = " ".join([slot for slot in data_dict[circuit][site] if slot[0] == "L"])
            new_row = pd.Series([site, circuit, slots, btb], index=["Site", "Circuit", "Slot", "BTB"])
            df = df.append(new_row, ignore_index=True)

    df.to_excel(f"Output-{sheet_value}.xlsx")


